# from msilib.schema import Error
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import *
from .forms import *
from authentication.forms import UploadFileForm
from django.contrib.auth.decorators import login_required, permission_required
from .upload_thread import *
import pandas as pd
from tablib import Dataset
from django.core.files.storage import FileSystemStorage
from appsystem.models import *
import os
from django.db.models import Sum
from accounting.models import Fiscal_year
from openpyxl import Workbook
from django.http import HttpResponse
from django.db.models import Sum ,Q,Count,F
from accounting.models import Fiscal_year
from inventory.models import Inventory
from .filters import *

@login_required(login_url='authentication:login')
@permission_required('fixedassets.custom_view_fixedassets')
def assets(request):
    app_model = Companymodule.objects.filter(tenant_id = request.user.devision.tenant_id.id)
    asset_list = FixedAsset.objects.all()
    user = request.user 
    myFilter = AssetsFilter(request.GET, queryset=asset_list)
    asset_list = myFilter.qs
    template = 'fixedassets/assets/asset.html'
    context = {
        'asset_list': asset_list,
        'heading': 'List of Assets',
        'pageview': 'Assets',
        'app_model':app_model,
        'myFilter':myFilter
    }
    return render(request,template,context)
    
@login_required(login_url='authentication:login')
@permission_required('fixedassets.custom_create_fixedassets')
def add_selectclassification(request):
    
    app_model = Companymodule.objects.filter(tenant_id = request.user.devision.tenant_id.id)
    if request.method == 'POST':
        form = AssetsClassificationForm(request.POST)
        if form.is_valid():
            classification = form.cleaned_data['classification']
            assetclassification = Classification.objects.get(name=classification)
            formclassification =assetclassification.name
            return redirect('fixedassets:new-asset', formclassification)
    else:
        form = AssetsClassificationForm()

    template = 'fixedassets/assets/classificationform.html'
    context = {
        'form':form,
        'heading': 'New Asset',
        'pageview': 'List of Assets',
        'app_model':app_model
    }
    return render(request,template,context)

@login_required(login_url='authentication:login')
@permission_required('fixedassets.custom_create_fixedassets')
def add_assets(request,formclassification):
    
    app_model = Companymodule.objects.filter(tenant_id = request.user.devision.tenant_id.id)
    if request.method == 'POST':
        if formclassification == 'Land':
            form = LandForm(request.POST,request=request)
        elif formclassification == 'Buldings And Other Structures':
            form = BuildingForm(request.POST,request=request)
        elif formclassification == 'Transport Equipments':
            form = TransportForm(request.POST,request=request)
        elif formclassification == 'Outdoor Machinery And Equipments':
            form = OutdoorForm(request.POST,request=request)
        elif formclassification == 'Indoor':
            form = IndoorForm(request.POST,request=request)
        else:
            form = WIPForm(request.POST,request=request)
        if form.is_valid():
            asset=form.save()
            inventory,created = Inventory.objects.get_or_create(product_id=asset.product,tenant_id = request.user.devision.tenant_id)
            inventory.avialable_quantity += 1
            inventory.save()
            messages.info(request,'Asset Saved')
            return redirect('fixedassets:detail-assets',asset.id)
            
    else:
        if formclassification == 'Land':
            form = LandForm(request=request)
        elif formclassification == 'Buldings And Other Structures':
            form = BuildingForm(request=request)
        elif formclassification == 'Transport Equipments':
            form = TransportForm(request=request)
        elif formclassification == 'Outdoor Machinery And Equipments':
            form = OutdoorForm(request=request)
        elif formclassification == 'Indoor':
            form = IndoorForm(request=request)
        else:
            form = WIPForm(request=request)
    if formclassification == 'Land':
        template = 'fixedassets/assets/create-land.html'
    elif formclassification == 'Buldings And Other Structures':
        template = 'fixedassets/assets/create-building.html'
    elif formclassification == 'Transport Equipments':
        template = 'fixedassets/assets/create-transport.html'
    elif formclassification == 'Outdoor Machinery And Equipments':
        template = 'fixedassets/assets/create-outdoor.html'
    elif formclassification == 'Indoor':
        template = 'fixedassets/assets/create-indoor.html'
    else:
        template = 'fixedassets/assets/create-wip.html'
    context = {
        'form':form,
        'heading': 'New Asset',
        'pageview': 'List of Assets',
        'app_model':app_model,
       
    }
    return render(request,template,context)


@login_required(login_url='authentication:login')
def load_subcategory(request):
    ipsascategory = request.GET.get('ipsascategory')
    subcategory = SubCategory.objects.filter(ipsascategory=ipsascategory).order_by('name')
    return render(request, 'authentication/district_dropdown_list.html', {'district': subcategory})

@login_required(login_url='authentication:login')
def load_user(request):
    subcostcenter = request.GET.get('subcostcenter')
    users = User.objects.filter(sub_division=subcostcenter).order_by('last_name')
    return render(request, 'authentication/user_dropdown_list.html', {'users': users})


@login_required(login_url='authentication:login')
@permission_required('fixedassets.custom_create_fixedassets')
def edit_assets(request,asset_id):
    
    app_model = Companymodule.objects.filter(tenant_id = request.user.devision.tenant_id.id)
    asset = FixedAsset.objects.get(id=asset_id)
    formclassification = asset.classification.name
    if request.method == 'POST':
        if formclassification == 'Land':
            form = LandForm(request.POST,request=request,instance=asset)
        elif formclassification == 'Buldings And Other Structures':
            form = BuildingForm(request.POST,request=request,instance=asset)
        elif formclassification == 'Transport Equipments':
            form = TransportForm(request.POST,request=request,instance=asset)
        elif formclassification == 'Outdoor Machinery And Equipments':
            form = OutdoorForm(request.POST,request=request,instance=asset)
        elif formclassification == 'Indoor':
            form = IndoorForm(request.POST,request=request,instance=asset)
        else:
            form = WIPForm(request.POST,request=request,instance=asset)
        if form.is_valid():
            form.save()
            messages.info(request,'Asset Saved')
            return redirect('fixedassets:detail-assets',asset.id)
            
    else:
        if formclassification == 'Land':
            form = LandForm(request=request,instance=asset)
        elif formclassification == 'Buldings And Other Structures':
            form = BuildingForm(request=request,instance=asset)
        elif formclassification == 'Transport Equipments':
            form = TransportForm(request=request,instance=asset)
        elif formclassification == 'Outdoor Machinery And Equipments':
            form = OutdoorForm(request=request,instance=asset)
        elif formclassification == 'Indoor':
            form = IndoorForm(request=request,instance=asset)
        else:
            form = WIPForm(request=request,instance=asset)
    if formclassification == 'Land':
        template = 'fixedassets/assets/create-land.html'
    elif formclassification == 'Buldings And Other Structures':
        template = 'fixedassets/assets/create-building.html'
    elif formclassification == 'Transport Equipments':
        template = 'fixedassets/assets/create-transport.html'
    elif formclassification == 'Outdoor Machinery And Equipments':
        template = 'fixedassets/assets/create-outdoor.html'

    elif formclassification == 'Indoor':
        template = 'fixedassets/assets/create-indoor.html'
    else:
        template = 'fixedassets/assets/create-wip.html'
    context = {
        'form':form,
        'heading': 'Update',
        'pageview': 'List of Assets',
        'app_model':app_model,
        'asset':asset
       
    }
    return render(request,template,context)

@login_required(login_url='authentication:login')
@permission_required('fixedassets.custom_create_fixedassets')
def assigned_assets(request,asset_id):
    app_model = Companymodule.objects.filter(tenant_id = request.user.devision.tenant_id.id)
    asset = FixedAsset.objects.get(id=asset_id)
    if request.method == 'POST':
        form = AssetsAssignmentForm(request.POST,request=request)
        if form.is_valid():
            assign = form.save(commit=False)
            assign.asset = asset
            assign.save()
            asset.costcenter = assign.costcenter
            asset.subcostcenter = assign.subcostcenter
            asset.location = assign.subcostcenter.location
            asset.user = assign.user
            asset.position = assign.user.grade
            asset.usagetype = assign.usagetype
            asset.status = 'Assigned'
            asset.currentstatus = 'In Use'
            asset.save()
            inventory = Inventory.objects.get(product_id=asset.product,tenant_id = request.user.devision.tenant_id)
            inventory.avialable_quantity -= 1
            inventory.save()
            
            messages.info(request,'Asset Assigned')
            return redirect('fixedassets:detail-assets',asset.id)
    else:
        form = AssetsAssignmentForm(request=request)
    template = 'fixedassets/assets/create-assign-asset.html'
    context = {
        'form':form,
        'heading': 'Assign An Asset',
        'pageview': 'List of Assets',
        'app_model':app_model,
        'asset':asset
       
    }
    return render(request,template,context)

@login_required(login_url='authentication:login')
@permission_required('fixedassets.custom_create_fixedassets')
def return_assets(request,assign_id):
    app_model = Companymodule.objects.filter(tenant_id = request.user.devision.tenant_id.id)
    assignment = FixedAssetsAssignment.objects.get(id=assign_id)
    asset = FixedAsset.objects.get(id=assignment.asset.id)
    if request.method == 'POST':
        form = AssetsReturnedForm(request.POST,)
        if form.is_valid():
            returndate = form.cleaned_data['returndate']
            assignment.returndate =returndate
            assignment.status = 'Returned'
            asset.status = 'Avialable'
            assignment.save()
            asset.save()
            inventory = Inventory.objects.get(product_id=asset.product,tenant_id = request.user.devision.tenant_id)
            inventory.avialable_quantity += 1
            inventory.save()
            messages.info(request,'Asset Returned')
            return redirect('fixedassets:detail-assets',asset.id)
    else:
        form = AssetsReturnedForm()
    template = 'fixedassets/assets/create-return-asset.html'
    context = {
        'form':form,
        'heading': 'Assign An Asset',
        'pageview': 'List of Assets',
        'app_model':app_model,
        'asset':asset
       
    }
    return render(request,template,context)


def delete_asset(request,asset_id):
    asset = FixedAsset.objects.get(id=asset_id)
    if asset.status == 'Avialable':
        inventory = Inventory.objects.get(product_id=asset.product,tenant_id = request.user.devision.tenant_id)
        inventory.avialable_quantity -= 1
        inventory.save()
    asset.delete()
    messages.error(request,'Asset Deleted')
    return redirect('fixedassets:asset-list')


@login_required(login_url='authentication:login')
@permission_required('fixedassets.custom_view_fixedassets')
def view_assets(request,asset_id):
    app_model = Companymodule.objects.filter(tenant_id = request.user.devision.tenant_id.id)
    asset = FixedAsset.objects.get(id=asset_id)
    assignment = FixedAssetsAssignment.objects.filter(asset=asset).order_by('-id')
    template = 'fixedassets/assets/view-asset.html'
    context = {
        
        'heading': 'Detail Asset',
        'pageview': 'List of Assets',
        'app_model':app_model,
        'asset':asset,
        'assignment':assignment,
       
    }
    return render(request,template,context)

@login_required(login_url='authentication:login')
@permission_required('fixedassets.custom_view_fixedassets')
def depreciation(request):
    app_model = Companymodule.objects.filter(tenant_id = request.user.devision.tenant_id.id)
    depreciation_list = Depreciation.objects.all().order_by('-id')
    current_depreciation = depreciation_list.first() 
    accumulated_depreciation = depreciation_list.aggregate(cc=Sum('depreciationvalue'))
    template = 'fixedassets/assets/depreciation/depreciation.html'
    context = {
        
        'heading': 'Depreciation',
        'pageview': 'Depreciation',
        'app_model':app_model,
        'depreciation_list':depreciation_list,
        'current_depreciation':current_depreciation,
        'accumulated_depreciation':accumulated_depreciation
    }
    return render(request,template,context)

@login_required(login_url='authentication:login')
@permission_required('fixedassets.custom_run_depreciation_on_fixed')
def run_depreciation(request):
    app_model = Companymodule.objects.filter(tenant_id = request.user.devision.tenant_id.id)
    
    if request.method == 'POST':
        form = DepreciationForm(request.POST)
        if form.is_valid():
            depreciation = form.save()
            assets = FixedAsset.objects.filter(depreciation='Yes',usefullifebalance__gt=0).exclude(currentstatus='Disposed')
            DepreciationThread(depreciation,assets).start()
            messages.info(request,'Depreciation Generation Stated')
            return redirect('fixedassets:depreciation-list')
    else:
        form = DepreciationForm()
    template = 'fixedassets/assets/depreciation/create-depreciation.html'
    context = {
        'form':form,
        'heading': 'Assign An Asset',
        'pageview': 'List of Assets',
        'app_model':app_model,
        
       
    }
    return render(request,template,context)


@login_required(login_url='authentication:login')
@permission_required('fixedassets.custom_view_fixedassets')
def view_depreciation(request,depreciation_id):
    app_model = Companymodule.objects.filter(tenant_id = request.user.devision.tenant_id.id)
    depreciation = Depreciation.objects.get(id=depreciation_id)
    depreciation_detail = DepreciationDetail.objects.filter(depreciation=depreciation)
    
    template = 'fixedassets/assets/depreciation/view-depreciation.html'
    context = {
        
        'heading': 'Current Depreciation',
        'pageview': 'Current Depreciation',
        'app_model':app_model,
        'depreciation':depreciation,
        'depreciation_detail':depreciation_detail,
       
    }
    return render(request,template,context)


@login_required(login_url='authentication:login')
@permission_required('fixedassets.custom_view_fixedassets')
def reevaluation(request):
    status = ['Retired','Disposed']
    app_model = Companymodule.objects.filter(tenant_id = request.user.devision.tenant_id.id)
    asset_list = Reevaluation.objects.all().order_by('-id')
    template = 'fixedassets/assets/reevaluation/due-evaluation.html'
    context = {
        
        'heading': 'Revaluated Assets',
        'pageview': 'List of Revaluated Assets',
        'app_model':app_model,
        'asset_list':asset_list, 
    }
    return render(request,template,context)

@login_required(login_url='authentication:login')
@permission_required('fixedassets.custom_run_depreciation_on_fixed')
def reevaluate_asset(request,asset_id):
    app_model = Companymodule.objects.filter(tenant_id = request.user.devision.tenant_id.id)
    asset = FixedAsset.objects.get(id=asset_id)
    if request.method == 'POST':
        form = ReevaluationForm(request.POST)
        if form.is_valid():
            evaluation = form.save(commit=False)
            evaluation.oldvalue = asset.value
            evaluation.previoususefullife = asset.usefullife
            evaluation.asset = asset
            evaluation.save()
            asset.value = form.cleaned_data['newvalue']
            asset.usefullife = form.cleaned_data['usefullife']
            asset.depreciatedlife = 0
            asset.accumulateddepreciation = 0.00
            asset.currentdepreciation = 0.00
            asset.save()
            messages.info(request,'Re-Evaluation Saved')
            return redirect('fixedassets:detail-assets',asset.id)
    else:
        form = ReevaluationForm()
    template = 'fixedassets/assets/reevaluation/create-evaluation.html'
    context = {
        'form':form,
        'heading': 'Revaluate Asset',
        'pageview': 'List of Assets',
        'app_model':app_model,
        
       
    }
    return render(request,template,context)


@login_required(login_url='authentication:login')
@permission_required('fixedassets.custom_create_fixedassets')
def disposal_asset(request,asset_id):
    app_model = Companymodule.objects.filter(tenant_id = request.user.devision.tenant_id.id)
    asset = FixedAsset.objects.get(id=asset_id)
    if request.method == 'POST':
        form = DesposalForm(request.POST)
        if form.is_valid():
            mdisposal=form.cleaned_data['methodofdesposal']
            proceedsfromsales = form.cleaned_data['methodofdesposal']
            disposal = form.save(commit=False)
            disposal.asset =asset
            disposal.save()
            asset.currentstatus = 'Disposed'
            asset.methodofdesposal=disposal.methodofdesposal
            if disposal.methodofdesposal == 'Auction' or disposal.methodofdesposal == 'Sales' or disposal.methodofdesposal == 'Scrapped':
                asset.proceedsfromsales = disposal.proceedsfromsales
            asset.desposal_date = disposal.desposal_date
            asset.save()
            if asset.status == 'Avialable':
                inventory = Inventory.objects.get(product_id=asset.product,tenant_id = request.user.devision.tenant_id)
                inventory.avialable_quantity -= 1
                inventory.save()
            messages.info(request,'Asset Disposed')
            return redirect('fixedassets:detail-assets',asset.id)
    else:
        form = DesposalForm()
    template = 'fixedassets/assets/disposal/create-disposal.html'
    context = {
        'form':form,
        'heading': 'Disposal of  Asset',
        'pageview': 'List of Assets',
        'app_model':app_model,
        
       
    }
    return render(request,template,context)

@login_required(login_url='authentication:login')
@permission_required('fixedassets.custom_view_fixedassets')
def disposal(request):
    app_model = Companymodule.objects.filter(tenant_id = request.user.devision.tenant_id.id)
    asset_list = Disposals.objects.all().order_by('-id')
    template = 'fixedassets/assets/disposal/disposal.html'
    context = {
        
        'heading': 'Disposal Of Assets',
        'pageview': 'List Of Disposal Of Assets',
        'app_model':app_model,
        'asset_list':asset_list, 
    }
    return render(request,template,context)


@login_required(login_url='authentication:login')
@permission_required('fixedassets.custom_create_fixedassets')
def report_selectclassification(request):
    
    app_model = Companymodule.objects.filter(tenant_id = request.user.devision.tenant_id.id)
    if request.method == 'POST':
        form = AssetsClassificationForm(request.POST)
        if form.is_valid():
            classification = form.cleaned_data['classification']
            assetclassification = Classification.objects.get(name=classification)
            

            return redirect('fixedassets:export-excel', assetclassification.id)
    else:
        form = AssetsClassificationForm()

    template = 'fixedassets/assets/classificationform.html'
    context = {
        'form':form,
        'heading': 'New Asset',
        'pageview': 'List of Assets',
        'app_model':app_model
    }
    return render(request,template,context)

def report_land(asset):
    wb = Workbook()
    ws = wb.active

    headers = ['Asset ID', 'Classification','Asset Description', 'Accounting Recognition','Amortisation','Usage/Domain','IPSAS Category','GFS Category','Sub Category','Size(Acre)',
    'Location/Town','Other Location/Town(Reg./Assemble/Town)','Ghana Post GPS Address(GT-0000-0000','Titled/Deed Number','Accounting Cost Center','User Dep/Section/Unit',
    'Staff ID','Full Name(First/Middle/Surname)','Position','Method of Acquisition','Current Status','Investment Property(Yes/No)',
    'Fund Source','Cost/Value GHC','Useful Life','Disposal Date (Day-Month-Year)','Method of Disposal','Proceeds Form Sale GHC','Comments']

    # Write headers to the first row of the worksheet
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

        # Write data from queryset to subsequent rows
    for row, obj in enumerate(asset, start=2):
        if obj.desposal_date:
            disposal_date = obj.desposal_date.strftime('%d-%b-%Y')
        else:
            disposal_date = obj.desposal_date

        ws.cell(row=row, column=1, value=obj.asset_id)
        ws.cell(row=row, column=2, value=obj.classification.name)
        ws.cell(row=row, column=3, value=obj.description)
        ws.cell(row=row, column=4, value=obj.accountingrecognition.name)
        ws.cell(row=row, column=5, value=obj.amotization)
        ws.cell(row=row, column=6, value=obj.usage)
        ws.cell(row=row, column=7, value=obj.ipsascategory.name)
        ws.cell(row=row, column=8, value=obj.gfscategory.name)
        ws.cell(row=row, column=9, value=obj.subcategory.name)
        ws.cell(row=row, column=10, value=obj.size)
        ws.cell(row=row, column=11, value=(str(obj.location.code)+ '-' + str(obj.location.location)))
        ws.cell(row=row, column=12, value=obj.subcostcenter.name)
        ws.cell(row=row, column=13, value=obj.ghanapostgpsaddress)
        ws.cell(row=row, column=14, value=obj.titled)
        ws.cell(row=row, column=15, value=(str(obj.costcenter.code) + ' - ' +str(obj.costcenter.name)))
        ws.cell(row=row, column=16, value=obj.subcostcenter.name)
        ws.cell(row=row, column=17, value=obj.user.staffid)
        ws.cell(row=row, column=18, value=(str(obj.user.first_name) + ' - ' + str(obj.user.last_name)))
        ws.cell(row=row, column=19, value=obj.position.name)
        ws.cell(row=row, column=20, value=obj.methodofacquisition.name)
        ws.cell(row=row, column=21, value=obj.currentstatus)
        ws.cell(row=row, column=22, value=obj.investmentproperty)
        ws.cell(row=row, column=23, value=(str(obj.fundsource.code) + ' - ' +str(obj.fundsource.funding)))
        ws.cell(row=row, column=24, value=obj.value)
        ws.cell(row=row, column=25, value=obj.usefullife)
        ws.cell(row=row, column=26, value=disposal_date)
        ws.cell(row=row, column=27, value=obj.methodofdesposal)
        ws.cell(row=row, column=28, value=obj.proceedsfromsales)
        ws.cell(row=row, column=29, value=obj.comments)

    # Create a response object with Excel MIME type
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=my_data.xlsx'  # Change the filename as needed

    # Save the workbook to the response
    wb.save(response)
    print(response)

    return response

def report_building(asset):
    wb = Workbook()
    ws = wb.active

    headers = ['Asset ID', 'Classification','Asset Description', 'Accounting Recognition','Depreciation','IPSAS Category','GFS Category','Sub Category','Unit/Qty',
    'Location/Town','Other Location/Town(Reg./Assemble/Town)','Ghana Post GPS Address(GT-0000-0000','Date Placed In Service','Accounting Cost Center','User Dep/Section/Unit','USage Type(Assign/Pool',
    'Staff ID','Full Name(First/Middle/Surname)','Position','Method of Acquisition','Current Status','Condition','Investment Property(Yes/No)',
    'Fund Source','Cost/Value GHC','Useful Life','Disposal Date (Day-Month-Year)','Method of Disposal','Proceeds Form Sale GHC','Comments','Current Depreciation GHC','Accumulated Depreciation GHC','Net Book Value GHC']

    # Write headers to the first row of the worksheet
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

        # Write data from queryset to subsequent rows
    for row, obj in enumerate(asset, start=2):
        if obj.desposal_date:
            disposal_date = obj.desposal_date.strftime('%d-%b-%Y')
        else:
            disposal_date = obj.desposal_date

        

        ws.cell(row=row, column=1, value=obj.asset_id)
        ws.cell(row=row, column=2, value=obj.classification.name)
        ws.cell(row=row, column=3, value=obj.description)
        ws.cell(row=row, column=4, value=obj.accountingrecognition.name)
        ws.cell(row=row, column=5, value=obj.depreciation)
        ws.cell(row=row, column=7, value=obj.ipsascategory.name)
        ws.cell(row=row, column=8, value=obj.gfscategory.name)
        ws.cell(row=row, column=9, value=obj.subcategory.name)
        ws.cell(row=row, column=10, value=obj.quantity)
        ws.cell(row=row, column=11, value=(str(obj.location.code)+ '-' + str(obj.location.location)))
        ws.cell(row=row, column=12, value=obj.subcostcenter.name)
        ws.cell(row=row, column=13, value=obj.ghanapostgpsaddress)
        ws.cell(row=row, column=14, value=dateplacedinservice)
        ws.cell(row=row, column=15, value=(str(obj.costcenter.code) + ' - ' +str(obj.costcenter.name)))
        ws.cell(row=row, column=16, value=obj.subcostcenter.name)
        ws.cell(row=row, column=16, value=obj.usagetype)
        ws.cell(row=row, column=17, value=obj.user.staffid)
        ws.cell(row=row, column=18, value=(str(obj.user.first_name) + ' - ' + str(obj.user.last_name)))
        ws.cell(row=row, column=19, value=obj.position.name)
        ws.cell(row=row, column=20, value=obj.methodofacquisition.name)
        ws.cell(row=row, column=21, value=obj.currentstatus)
        ws.cell(row=row, column=21, value=obj.conditions)
        ws.cell(row=row, column=22, value=obj.investmentproperty)
        ws.cell(row=row, column=23, value=(str(obj.fundsource.code) + ' - ' +str(obj.fundsource.funding)))
        ws.cell(row=row, column=24, value=obj.value)
        ws.cell(row=row, column=25, value=obj.usefullife)
        ws.cell(row=row, column=26, value=disposal_date)
        ws.cell(row=row, column=27, value=obj.methodofdesposal)
        ws.cell(row=row, column=28, value=obj.proceedsfromsales)
        ws.cell(row=row, column=29, value=obj.comments)
        ws.cell(row=row, column=30, value=obj.currentdepreciation)
        ws.cell(row=row, column=31, value=obj.accumulateddepreciation)
        ws.cell(row=row, column=32, value=obj.netbookvalue)

    # Create a response object with Excel MIME type
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=my_data.xlsx'  # Change the filename as needed

    # Save the workbook to the response
    wb.save(response)
    print(response)
     
    return response
    
def transport(asset):
    wb = Workbook()
    ws = wb.active

    headers = ['Asset ID', 'Classification','Asset Description', 'Registration Number','Accounting Recognition','Depreciation','IPSAS Category','GFS Category','Sub Category','Unit/Qty',
    'Location/Town','Other Location/Town(Reg./Assemble/Town)','Date Placed In Service','Colour','Chassis Number','Engine Serial Number','Manufacturer Name','Model Number/Name','Model Year','Accounting Cost Center','User Dep/Section/Unit','USage Type(Assign/Pool',
    'Staff ID','Full Name(First/Middle/Surname)','Position','Method of Acquisition','Current Status','Condition','Investment Property(Yes/No)',
    'Fund Source','Cost/Value GHC','Useful Life','Disposal Date (Day-Month-Year)','Method of Disposal','Proceeds Form Sale GHC','Comments','Current Depreciation GHC','Accumulated Depreciation GHC','Net Book Value GHC']

    # Write headers to the first row of the worksheet
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

        # Write data from queryset to subsequent rows
    for row, obj in enumerate(asset, start=2):
        if obj.desposal_date:
            disposal_date = obj.desposal_date.strftime('%d-%b-%Y')
        else:
            disposal_date = obj.desposal_date
        
        if obj.dateplacedinservice:
            dateplacedinservice = obj.dateplacedinservice.strftime('%d-%b-%Y')
        else:
            dateplacedinservice = obj.dateplacedinservice

        ws.cell(row=row, column=1, value=obj.asset_id)
        ws.cell(row=row, column=2, value=obj.classification.name)
        ws.cell(row=row, column=3, value=obj.description)
        ws.cell(row=row, column=4, value=obj.registrationnumber)
        ws.cell(row=row, column=5, value=obj.accountingrecognition.name)
        ws.cell(row=row, column=6, value=obj.depreciation)
        ws.cell(row=row, column=7, value=obj.ipsascategory.name)
        ws.cell(row=row, column=8, value=obj.gfscategory.name)
        ws.cell(row=row, column=9, value=obj.subcategory.name)
        ws.cell(row=row, column=10, value=obj.quantity)
        ws.cell(row=row, column=11, value=(str(obj.location.code)+ '-' + str(obj.location.location)))
        ws.cell(row=row, column=12, value=obj.subcostcenter.name)
        ws.cell(row=row, column=13, value=dateplacedinservice)
        ws.cell(row=row, column=14, value=obj.colour)
        ws.cell(row=row, column=15, value=obj.chassisno)
        ws.cell(row=row, column=16, value=obj.engineserialno)
        ws.cell(row=row, column=17, value=obj.manufacturer.name)
        ws.cell(row=row, column=18, value=obj.model)
        ws.cell(row=row, column=19, value=obj.modelyear)
        ws.cell(row=row, column=20, value=(str(obj.costcenter.code) + ' - ' +str(obj.costcenter.name)))
        ws.cell(row=row, column=21, value=obj.subcostcenter.name)
        ws.cell(row=row, column=22, value=obj.usagetype)
        ws.cell(row=row, column=23, value=obj.user.staffid)
        ws.cell(row=row, column=24, value=(str(obj.user.first_name) + ' - ' + str(obj.user.last_name)))
        ws.cell(row=row, column=25, value=obj.position.name)
        ws.cell(row=row, column=26, value=obj.methodofacquisition.name)
        ws.cell(row=row, column=27, value=obj.currentstatus)
        ws.cell(row=row, column=28, value=obj.conditions)
        ws.cell(row=row, column=29, value=obj.investmentproperty)
        ws.cell(row=row, column=30, value=(str(obj.fundsource.code) + ' - ' +str(obj.fundsource.funding)))
        ws.cell(row=row, column=31, value=obj.value)
        ws.cell(row=row, column=32, value=obj.usefullife)
        ws.cell(row=row, column=33, value=disposal_date)
        ws.cell(row=row, column=34, value=obj.methodofdesposal)
        ws.cell(row=row, column=35, value=obj.proceedsfromsales)
        ws.cell(row=row, column=36, value=obj.comments)
        ws.cell(row=row, column=37, value=obj.currentdepreciation)
        ws.cell(row=row, column=38, value=obj.accumulateddepreciation)
        ws.cell(row=row, column=39, value=obj.netbookvalue)

    # Create a response object with Excel MIME type
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=my_data.xlsx'  # Change the filename as needed

    # Save the workbook to the response
    wb.save(response)
    print(response)
     
    return response

def outdoor(asset):
    wb = Workbook()
    ws = wb.active

    headers = ['Asset ID', 'Classification','Asset Description','Accounting Recognition','Depreciation','IPSAS Category','GFS Category','Sub Category','Unit/Qty',
    'Location/Town','Other Location/Town(Reg./Assemble/Town)','Ghana Post GPS Address(GT-0000-0000','Date Placed In Service','Chassis Number','Engine Serial Number','Manufacturer Name','Model Number/Name','Model Year','Tag Number','Accounting Cost Center','User Dep/Section/Unit','USage Type(Assign/Pool',
    'Staff ID','Full Name(First/Middle/Surname)','Position','Method of Acquisition','Current Status','Condition','Investment Property(Yes/No)',
    'Fund Source','Cost/Value GHC','Useful Life','Disposal Date (Day-Month-Year)','Method of Disposal','Proceeds Form Sale GHC','Comments','Current Depreciation GHC','Accumulated Depreciation GHC','Net Book Value GHC']

    # Write headers to the first row of the worksheet
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

        # Write data from queryset to subsequent rows
    for row, obj in enumerate(asset, start=2):
        if obj.desposal_date:
            disposal_date = obj.desposal_date.strftime('%d-%b-%Y')
        else:
            disposal_date = obj.desposal_date

        if obj.dateplacedinservice:
            dateplacedinservice = obj.dateplacedinservice.strftime('%d-%b-%Y')
        else:
            dateplacedinservice = obj.dateplacedinservice


        ws.cell(row=row, column=1, value=obj.asset_id)
        ws.cell(row=row, column=2, value=obj.classification.name)
        ws.cell(row=row, column=3, value=obj.description)
        ws.cell(row=row, column=4, value=obj.accountingrecognition.name)
        ws.cell(row=row, column=5, value=obj.depreciation)
        ws.cell(row=row, column=6, value=obj.ipsascategory.name)
        ws.cell(row=row, column=7, value=obj.gfscategory.name)
        ws.cell(row=row, column=8, value=obj.subcategory.name)
        ws.cell(row=row, column=9, value=obj.quantity)
        ws.cell(row=row, column=10, value=(str(obj.location.code)+ '-' + str(obj.location.location)))
        ws.cell(row=row, column=11, value=obj.subcostcenter.name)
        ws.cell(row=row, column=12, value=obj.ghanapostgpsaddress)
        ws.cell(row=row, column=13, value=dateplacedinservice)
        ws.cell(row=row, column=14, value=obj.chassisno)
        ws.cell(row=row, column=15, value=obj.engineserialno)
        ws.cell(row=row, column=16, value=obj.manufacturer.name)
        ws.cell(row=row, column=17, value=obj.model)
        ws.cell(row=row, column=18, value=obj.modelyear)
        ws.cell(row=row, column=19, value=obj.tagno)
        ws.cell(row=row, column=20, value=(str(obj.costcenter.code) + ' - ' +str(obj.costcenter.name)))
        ws.cell(row=row, column=21, value=obj.subcostcenter.name)
        ws.cell(row=row, column=22, value=obj.usagetype)
        ws.cell(row=row, column=23, value=obj.user.staffid)
        ws.cell(row=row, column=24, value=(str(obj.user.first_name) + ' - ' + str(obj.user.last_name)))
        ws.cell(row=row, column=25, value=obj.position.name)
        ws.cell(row=row, column=26, value=obj.methodofacquisition.name)
        ws.cell(row=row, column=27, value=obj.currentstatus)
        ws.cell(row=row, column=28, value=obj.conditions)
        ws.cell(row=row, column=29, value=obj.investmentproperty)
        ws.cell(row=row, column=30, value=(str(obj.fundsource.code) + ' - ' +str(obj.fundsource.funding)))
        ws.cell(row=row, column=31, value=obj.value)
        ws.cell(row=row, column=32, value=obj.usefullife)
        ws.cell(row=row, column=33, value=disposal_date)
        ws.cell(row=row, column=34, value=obj.methodofdesposal)
        ws.cell(row=row, column=35, value=obj.proceedsfromsales)
        ws.cell(row=row, column=36, value=obj.comments)
        ws.cell(row=row, column=37, value=obj.currentdepreciation)
        ws.cell(row=row, column=38, value=obj.accumulateddepreciation)
        ws.cell(row=row, column=39, value=obj.netbookvalue)

    # Create a response object with Excel MIME type
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=my_data.xlsx'  # Change the filename as needed

    # Save the workbook to the response
    wb.save(response)
    print(response)
     
    return response

def indoor(asset):
    wb = Workbook()
    ws = wb.active

    headers = ['Asset ID', 'Classification','Asset Description','Accounting Recognition','Depreciation','IPSAS Category','GFS Category','Sub Category','Unit/Qty',
    'Location/Town','Other Location/Town(Reg./Assemble/Town)','Date Placed In Service','Serial Number','Manufacturer Name','Tag Number','Accounting Cost Center','User Dep/Section/Unit','USage Type(Assign/Pool',
    'Staff ID','Full Name(First/Middle/Surname)','Position','Method of Acquisition','Current Status','Condition','Investment Property(Yes/No)',
    'Fund Source','Cost/Value GHC','Useful Life','Disposal Date (Day-Month-Year)','Method of Disposal','Proceeds Form Sale GHC','Comments','Current Depreciation GHC','Accumulated Depreciation GHC','Net Book Value GHC']

    # Write headers to the first row of the worksheet
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

        # Write data from queryset to subsequent rows
    for row, obj in enumerate(asset, start=2):
        if obj.desposal_date:
            disposal_date = obj.desposal_date.strftime('%d-%b-%Y')
        else:
            disposal_date = obj.desposal_date

        if obj.dateplacedinservice:
            dateplacedinservice = obj.dateplacedinservice.strftime('%d-%b-%Y')
        else:
            dateplacedinservice = obj.dateplacedinservice



        ws.cell(row=row, column=1, value=obj.asset_id)
        ws.cell(row=row, column=2, value=obj.classification.name)
        ws.cell(row=row, column=3, value=obj.description)
        ws.cell(row=row, column=4, value=obj.accountingrecognition.name)
        ws.cell(row=row, column=5, value=obj.depreciation)
        ws.cell(row=row, column=6, value=obj.ipsascategory.name)
        ws.cell(row=row, column=7, value=obj.gfscategory.name)
        ws.cell(row=row, column=8, value=obj.subcategory.name)
        ws.cell(row=row, column=9, value=obj.quantity)
        ws.cell(row=row, column=10, value=(str(obj.location.code)+ '-' + str(obj.location.location)))
        ws.cell(row=row, column=11, value=obj.subcostcenter.name)
        ws.cell(row=row, column=12, value=dateplacedinservice)
        ws.cell(row=row, column=13, value=obj.chassisno)
        ws.cell(row=row, column=14, value=obj.manufacturer.name)
        ws.cell(row=row, column=15, value=obj.tagno)
        ws.cell(row=row, column=16, value=(str(obj.costcenter.code) + ' - ' +str(obj.costcenter.name)))
        ws.cell(row=row, column=17, value=obj.subcostcenter.name)
        ws.cell(row=row, column=18, value=obj.usagetype)
        ws.cell(row=row, column=19, value=obj.user.staffid)
        ws.cell(row=row, column=20, value=(str(obj.user.first_name) + ' - ' + str(obj.user.last_name)))
        ws.cell(row=row, column=21, value=obj.position.name)
        ws.cell(row=row, column=22, value=obj.methodofacquisition.name)
        ws.cell(row=row, column=23, value=obj.currentstatus)
        ws.cell(row=row, column=24, value=obj.conditions)
        ws.cell(row=row, column=25, value=obj.investmentproperty)
        ws.cell(row=row, column=26, value=(str(obj.fundsource.code) + ' - ' +str(obj.fundsource.funding)))
        ws.cell(row=row, column=27, value=obj.value)
        ws.cell(row=row, column=28, value=obj.usefullife)
        ws.cell(row=row, column=29, value=disposal_date)
        ws.cell(row=row, column=30, value=obj.methodofdesposal)
        ws.cell(row=row, column=31, value=obj.proceedsfromsales)
        ws.cell(row=row, column=32, value=obj.comments)
        ws.cell(row=row, column=33, value=obj.currentdepreciation)
        ws.cell(row=row, column=34, value=obj.accumulateddepreciation)
        ws.cell(row=row, column=35, value=obj.netbookvalue)

    # Create a response object with Excel MIME type
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=my_data.xlsx'  # Change the filename as needed

    # Save the workbook to the response
    wb.save(response)
    print(response)
     
    return response

def wip(asset):
    wb = Workbook()
    ws = wb.active

    headers = ['Asset ID', 'Classification','Asset Description','Accounting Recognition','Depreciation','IPSAS Category','GFS Category','Unit/Qty',
    'Location/Town','Other Location/Town(Reg./Assemble/Town)','Ghana Post GPS Address(GT-0000-0000','Commencement Date(Day-Month-Year)','Expected Completion Date(Day-Month-Year)','Accounting Cost Center',
    'Staff ID','Full Name(First/Middle/Surname)','Position','Accounting Status','Current Status','Fund Source','Cost B/F GHC','Current Period Cost GHC','Disposal Date (Day-Month-Year)','Method of Disposal','Proceeds Form Sale GHC',]

    # Write headers to the first row of the worksheet
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

        # Write data from queryset to subsequent rows
    for row, obj in enumerate(asset, start=2):
        if obj.desposal_date:
            disposal_date = obj.desposal_date.strftime('%d-%b-%Y')
        else:
            disposal_date = obj.desposal_date

        if obj.commencement_date:
            commencement_date = obj.commencement_date.strftime('%d-%b-%Y')
        else:
            commencement_date = obj.commencement_date
        
        if obj.expectedcompletion_date:
            expectedcompletion_date = obj.expectedcompletion_date.strftime('%d-%b-%Y')
        else:
            expectedcompletion_date = obj.expectedcompletion_date


        ws.cell(row=row, column=1, value=obj.asset_id)
        ws.cell(row=row, column=2, value=obj.classification.name)
        ws.cell(row=row, column=3, value=obj.description)
        ws.cell(row=row, column=4, value=obj.accountingrecognition.name)
        ws.cell(row=row, column=5, value=obj.depreciation)
        ws.cell(row=row, column=6, value=obj.ipsascategory.name)
        ws.cell(row=row, column=7, value=obj.gfscategory.name)
        ws.cell(row=row, column=8, value=obj.quantity)
        ws.cell(row=row, column=9, value=(str(obj.location.code)+ '-' + str(obj.location.location)))
        ws.cell(row=row, column=10, value=obj.subcostcenter.name)
        ws.cell(row=row, column=11, value=obj.ghanapostgpsaddress)
        ws.cell(row=row, column=12, value=commencement_date)
        ws.cell(row=row, column=13, value=expectedcompletion_date)
        ws.cell(row=row, column=14, value=(str(obj.costcenter.code) + ' - ' +str(obj.costcenter.name)))
        ws.cell(row=row, column=15, value=obj.user.staffid)
        ws.cell(row=row, column=16, value=(str(obj.user.first_name) + ' - ' + str(obj.user.last_name)))
        ws.cell(row=row, column=17, value=obj.position.name)
        ws.cell(row=row, column=18, value=obj.accountingstatus)
        ws.cell(row=row, column=19, value=obj.currentstatus)
        ws.cell(row=row, column=20, value=(str(obj.fundsource.code) + ' - ' +str(obj.fundsource.funding))) 
        ws.cell(row=row, column=21, value=obj.costbf)
        ws.cell(row=row, column=22, value=obj.currentperiodcost)
        ws.cell(row=row, column=23, value=obj.costcf)
        ws.cell(row=row, column=24, value=disposal_date)
        ws.cell(row=row, column=25, value=obj.methodofdesposal)
        ws.cell(row=row, column=26, value=obj.proceedsfromsales)
        ws.cell(row=row, column=27, value=obj.comments)
       

    # Create a response object with Excel MIME type
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=my_data.xlsx'  # Change the filename as needed

    # Save the workbook to the response
    wb.save(response)
    print(response)
     
    return response


def export_excel (request,assetclassification):
    classification = Classification.objects.get(id=assetclassification)
    asset = FixedAsset.objects.filter(classification = assetclassification)
    
    if classification.name == 'Land':
        response =report_land(asset)
    elif classification.name =='Buldings And Other Structures':
       response = report_building(asset)
    elif classification.name =='Transport Equipments':
           response = transport(asset)
    elif classification.name == 'Outdoor Machinery And Equipments':
        response = outdoor(asset)
    elif classification.name == 'Indoor':
        response = indoor(asset)
    else:
        response = wip(asset)

    return response 



def assetdashboard(request):
    app_model = Companymodule.objects.filter(tenant_id = request.user.devision.tenant_id.id)
    exclude_list = ['Disposed','Retired']
    costcenter = Devision.objects.all()
    classification = Classification.objects.all()
    depreciation_list = Depreciation.objects.all().order_by('-id')
    current_depreciation = depreciation_list.first() 
    accumulated_depreciation = depreciation_list.aggregate(total=Sum('depreciationvalue'))
    fixedasset_list = FixedAsset.objects.all().exclude(currentstatus='Disposed')
    fixedasset = FixedAsset.objects.all()
    total_fixedasset_status = fixedasset.values('currentstatus').annotate(total=Count('id'))
    total_fixedasset_netbook = fixedasset_list.values('currentstatus').annotate(netbookvalue=Sum('netbookvalue')).values('currentstatus','netbookvalue').exclude(netbookvalue__lte=0)
    total_fixedasset_costcenter = costcenter.values('name').annotate(total=Count('assetcostcenter__id'),netbookvalue=Sum('assetcostcenter__netbookvalue')).values('name','total','netbookvalue')
    # total_fixedasset_classification = classification.values('name').annotate(total=Count('classification__id'),netbookvalue=Sum('classification__netbookvalue')).values('name','total','netbookvalue').exclude(total__lte=0)
    total_netbook_value = fixedasset_list.aggregate(total=Sum('netbookvalue'))
    total_asset_disposed =  Disposals.objects.all().aggregate(total=Count('id'))
    total_asset_value_disposed =  Disposals.objects.all().aggregate(total=Sum('proceedsfromsales'))
    current_disposal_value=Disposals.objects.filter(accountingyear__status='Active').aggregate(total=Sum('proceedsfromsales'))
    total_asset_value_evaluated =  Reevaluation.objects.all().aggregate(total=Sum('newvalue'))
    total_asset_evaluated =  Reevaluation.objects.all().aggregate(total=Count('id'))
    current_evaluation_value=Reevaluation.objects.filter(accountingyear__status='Active').aggregate(total=Sum('newvalue'))
    total_fixedasset_classification =(
        FixedAsset.objects.exclude(currentstatus='Disposed')
        .values('classification__name')  # Grouping by classification name
        .annotate(
            total_assets=Count('id'),  # Counting asset_id for each group
            total_netbook_value=Sum('netbookvalue')  # Summing netbookvalue for each group
        ).exclude(total_netbook_value__lte=0.00)
        )
    template = 'dashboard/fixedasset.html'
    context = {
        
        'heading': 'Dashboard',
        'pageview': 'Dashboard',
        'app_model':app_model,
        'current_depreciation':current_depreciation, 
        'accumulated_depreciation':accumulated_depreciation,
        'total_fixedasset_status':total_fixedasset_status,
        'total_fixedasset_netbook':total_fixedasset_netbook,
        'total_fixedasset_costcenter':total_fixedasset_costcenter,
        'total_fixedasset_classification':total_fixedasset_classification,
        'total_netbook_value':total_netbook_value,
        'total_asset_disposed':total_asset_disposed,
        'total_asset_value_disposed':total_asset_value_disposed,
        'current_disposal_value':current_disposal_value,
        'total_asset_value_evaluated':total_asset_value_evaluated,
        'total_asset_evaluated':total_asset_evaluated,
        'current_evaluation_value':current_evaluation_value,
       
       


    }
    return render(request,template,context)